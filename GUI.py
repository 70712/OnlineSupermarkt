from openpyxl import load_workbook
from database import *
from product import *
from PDF import *
from tkinter import filedialog




##############################
# Grundlegende Variablen
wb = load_workbook('products.xlsx')
sheet = wb['Products']

client = pymongo.MongoClient('localhost', 27017)
database = client["database"]
db = Database(database)
db.delete_many()

list_products = []  # Instanzen der Produkte

color_grey = '#CCCCCC'
color_lightgrey = '#EAEAEA'




##############################
# Insert Products into Datebase
for i in range(2, 78):
    values = {}
    for j in range(1, 13):
        values[j] = sheet.cell(row=i, column=j).value
    if values[10] == 'True':
        values[10] = True
    else:
        values[10] = False
    values[11] = 'products\\' + str(values[11])
    values[12] = 'products\\' + str(values[12])
    db.insert(int(values[1]), str(values[2]), str(values[3]), float(values[4]), float(values[5]), str(values[6]), int(values[7]), int(values[8]), float(values[9]), values[10], values[11], values[12])




##############################
# Hauptfenster
window = Tk()
window.title("Supermarkt")
window.resizable(0, 0)




##############################
# Hauptfenster-Images
img_hotbar = PhotoImage(file="hotbar.gif", master=window)
img_search = PhotoImage(file="search.gif", master=window)
img_account = PhotoImage(file="account.gif", master=window)
img_shoppingbag = PhotoImage(file="shoppingbag.gif", master=window)
img_shoppingbag_empty = PhotoImage(file="shoppingbag_empty.gif", master=window)
img_shoppingbag_full = PhotoImage(file="shoppingbag_full.gif", master=window)

img_angebot = PhotoImage(file="angebot.gif", master=window)
img_blank = PhotoImage(file="blank.gif", master=window)
img_line = PhotoImage(file="line_small.gif", master=window)
img_previous = PhotoImage(file="left.gif", master=window)
img_next = PhotoImage(file="right.gif", master=window)

img_filter = PhotoImage(file="filter.gif", master=window)
img_filter_sortierennach = PhotoImage(file="filter_sortierennach.gif", master=window)
img_filter_preis = PhotoImage(file="filter_preis.gif", master=window)
img_filter_kategorie = PhotoImage(file="filter_kategorie.gif", master=window)
img_filter_reduziert = PhotoImage(file="filter_reduziert.gif", master=window)
img_filter_verfuegbar = PhotoImage(file="filter_verfügbar.gif", master=window)
img_filter_bio = PhotoImage(file="filter_bio.gif", master=window)
img_filter_hersteller = PhotoImage(file="filter_hersteller.gif", master=window)
img_filter_reset = PhotoImage(file="reset.gif", master=window)




##############################
# Hauptframe
frame = Frame(window, width=1280, height=800, bg='white')




##############################
# Hotbar (Logo, Suchbar, Account und Warenkorb)
hotbar = Frame(frame, width=1280, height=124, bg='white')
hotbar.grid(rows=1, columns=1)

var_searchinput = StringVar(master=window, value='')
var_searchnotfound = StringVar(master=window, value='')
var_searchresults = ListVar([])

# Hintergrund
hotbar_background = Label(hotbar, image=img_hotbar, bg='white', borderwidth=0)
hotbar_background.photo = img_hotbar
hotbar_background.place(x=0, y=0)

shoppingbagbutton = Button(hotbar, image=img_shoppingbag_empty, bg='white', borderwidth=0)
shoppingbagbutton.photo = img_shoppingbag_empty
var_shoppingbag = StringVar(master=window, value=0)
shoppingbag_instance = Shoppingbag(db, shoppingbagbutton, var_shoppingbag, img_shoppingbag_empty, img_shoppingbag_full)
shoppingbag_instance.update()

##############################
# Angebot
angebot = Frame(frame, width=1280, height=166, bg='white')
angebot.grid(rows=2, columns=1)




##############################
# Filter
filter = Frame(frame, width=1280, height=80, bg='white')
filter.grid(rows=3, columns=1)

# Variablen
var_sortierennach = StringVar(master=window, value=0)
var_0_1 = IntVar(master=window, value=0)
var_1_2 = IntVar(master=window, value=0)
var_2_5 = IntVar(master=window, value=0)
var_5_inf = IntVar(master=window, value=0)
var_backwaren = IntVar(master=window, value=0)
var_drogerieartikel = IntVar(master=window, value=0)
var_eierundmilchprodukte = IntVar(master=window, value=0)
var_fischfleischundwurst = IntVar(master=window, value=0)
var_gemuese = IntVar(master=window, value=0)
var_getraenke = IntVar(master=window, value=0)
var_obst = IntVar(master=window, value=0)
var_suesswaren = IntVar(master=window, value=0)
var_teigwaren = IntVar(master=window, value=0)
var_tiefkuehlprodukte = IntVar(master=window, value=0)
var_reduziert = IntVar(master=window, value=0)
var_verfuegbar = IntVar(master=window, value=0)
var_bio = IntVar(master=window, value=0)


##############################
# Produktübersicht
def on_mousewheel(event):
    overview_canvas.yview_scroll(-1 * int(event.delta / 120), "units")

overview = Frame(frame, width=1280, height=430, bg='white')
overview.grid(rows=4, columns=1)
overview.grid_rowconfigure(0, weight=1)
overview.grid_columnconfigure(0, weight=1)
overview.grid_propagate(False)

overview_canvas = Canvas(overview, width=1265, height=430, bg="white")
overview_canvas.grid(row=0, column=0)
overview_canvas.bind_all("<MouseWheel>", on_mousewheel)

overview_scrollbar = Scrollbar(overview, orient="vertical", command=overview_canvas.yview)
overview_scrollbar.grid(row=0, column=1, sticky='ns')

overview_canvas.configure(yscrollcommand=overview_scrollbar.set)

overview_frame = Frame(overview_canvas, bg="white")
overview_frame.grid(row=0, column=0)

overview_canvas.create_window((0, 0), window=overview_frame, anchor='nw')


# Suche starten
def searchbutton_clicked():
    window.focus()
    searchinput = str(var_searchinput.get())
    clearOverview()
    list = []

    if searchinput == '':
        list = db.artikelnummern_all
    else:
        query = {'name': {'$regex': searchinput, '$options': 'i'}}  # Regulärer Ausdruck: beinhalte searchinput und ignoriere Groß- und Kleinschreibung
        doc = db.col.find(query)
        for entry in doc:
            list.append(int(entry['artikelnummer']))
    if len(list) == 0:
        var_searchnotfound.set('Leider nichts gefunden!')
    else:
        var_searchnotfound.set('')
    var_searchresults.set(list)
    showProducts()

# Filter Funktionen

def sort(products, case):
    dictionary = {}
    output = []
    for artikelnummer in products:
        query = {'artikelnummer': artikelnummer}
        doc = db.col.find(query)
        entry = None
        for e in doc:
            entry = e
        if case == 0:   # Alphabetisch (A-Z)
            dictionary[artikelnummer] = str(entry['name'])
            output = sorted(dictionary.items(), key=lambda kv: (kv[1], kv[0]))
        elif case == 1: # Alphabetisch (Z-A)
            dictionary[artikelnummer] = str(entry['name'])
            output = sorted(dictionary.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
        elif case == 2: # Beliebtheit
            dictionary[artikelnummer] = int(entry['verkauft'])
            output = sorted(dictionary.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
        elif case == 3: # Preis absteigend
            dictionary[artikelnummer] = float(entry['preis'] - entry['rabatt'])
            output = sorted(dictionary.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
        elif case == 4: # Preis aufsteigend
            dictionary[artikelnummer] = float(entry['preis'] - entry['rabatt'])
            output = sorted(dictionary.items(), key=lambda kv: (kv[1], kv[0]))
    if case == 0:   # Alphabetisch (A-Z)
        output = sorted(dictionary.items(), key=lambda kv: (kv[1], kv[0]))
    elif case == 1: # Alphabetisch (Z-A)
        output = sorted(dictionary.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
    elif case == 2: # Beliebtheit
        output = sorted(dictionary.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
    elif case == 3: # Preis absteigend
        output = sorted(dictionary.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
    elif case == 4: # Preis aufsteigend
        output = sorted(dictionary.items(), key=lambda kv: (kv[1], kv[0]))
    return list(dict(output).keys())

def preis():
    col = db.col
    output = []
    if int(var_0_1.get()) == 0 and int(var_1_2.get()) == 0 and int(var_2_5.get()) == 0 and int(var_5_inf.get()) == 0:
        return db.artikelnummern_all
    if int(var_0_1.get()) == 1:
        query = {'preis': {'$lt': 1}}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])
    if int(var_1_2.get()) == 1:
        query = {'preis': {'$gte': 1, '$lt': 2}}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])
    if int(var_2_5.get()) == 1:
        query = {'preis': {'$gte': 2, '$lt': 5}}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])
    if int(var_5_inf.get()) == 1:
        query = {'preis': {'$gte': 5}}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])
    return output

def kategorie():
    col = db.col
    output = []
    if int(var_backwaren.get()) == 0 and int(var_drogerieartikel.get()) == 0 and int(var_eierundmilchprodukte.get()) == 0 and int(var_fischfleischundwurst.get()) == 0 and int(var_gemuese.get()) == 0 and int(var_getraenke.get()) == 0 and int(var_obst.get()) == 0 and int(var_suesswaren.get()) == 0 and int(var_teigwaren.get()) == 0 and int(var_tiefkuehlprodukte.get()) == 0:
        return db.artikelnummern_all

    if int(var_backwaren.get()) == 1:
        query = {'kategorie': 'Backwaren'}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])

    if int(var_drogerieartikel.get()) == 1:
        query = {'kategorie': 'Drogerieartikel'}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])

    if int(var_eierundmilchprodukte.get()) == 1:
        query = {'kategorie': 'Eier und Milchprodukte'}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])

    if int(var_fischfleischundwurst.get()) == 1:
        query = {'kategorie': 'Fisch, Fleisch und Wurst'}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])

    if int(var_gemuese.get()) == 1:
        query = {'kategorie': 'Gemüse'}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])

    if int(var_getraenke.get()) == 1:
        query = {'kategorie': 'Getränke'}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])

    if int(var_obst.get()) == 1:
        query = {'kategorie': 'Obst'}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])

    if int(var_suesswaren.get()) == 1:
        query = {'kategorie': 'Süßwaren'}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])

    if int(var_teigwaren.get()) == 1:
        query = {'kategorie': 'Teigwaren'}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])

    if int(var_tiefkuehlprodukte.get()) == 1:
        query = {'kategorie': 'Tiefkühlprodukte'}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])
    return output

def reduziert():
    col = db.col
    output = []
    if int(var_reduziert.get()) == 0:
        return db.artikelnummern_all

    if int(var_reduziert.get()) == 1:
        query = {'rabatt': {'$gt': 0}}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])
    return output

def verfuegbar():
    col = db.col
    output = []
    if int(var_verfuegbar.get()) == 0:
        return db.artikelnummern_all

    if int(var_verfuegbar.get()) == 1:
        query = {'auf_lager': {'$gt': 0}}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])
    return output

def bio():
    col = db.col
    output = []
    if int(var_bio.get()) == 0:
        return db.artikelnummern_all

    if int(var_bio.get()) == 1:
        query = {'bio': True}
        doc = col.find(query)
        for entry in doc:
            output.append(entry['artikelnummer'])
    return output

##############################
# Produkte
def clearOverview():
    for prod in list_products:
        overview_canvas.delete(prod.getid())
    overview_canvas.config(scrollregion=overview_canvas.bbox("all"))

def createProducts(products):
    x, y = 3, 5
    products = sort(products, int(var_sortierennach.get()))
    for artikelnummer in products:
        prod = Product(window, x, y, overview_canvas, db, artikelnummer, shoppingbag_instance)
        list_products.append(prod)
        x += 253
        if x == 1268:
            x = 3
            y += 430
    overview_canvas.config(scrollregion=overview_canvas.bbox("all"))

createProducts(db.artikelnummern_all)

def showProducts():
    x, y = 3, 5
    clearOverview()

    products = []

    # Preis filtern
    list_preis = preis()
    if len(products) > 0:
        products = list(set(products).intersection(list_preis))
    else:
        products = list_preis

    # Kategorie filtern
    list_kategorie = kategorie()
    if len(products) > 0:
        products = list(set(products).intersection(list_kategorie))
    else:
        products = list_kategorie

    # Reduziert filtern
    list_reduziert = reduziert()
    if len(products) > 0:
        products = list(set(products).intersection(list_reduziert))
    else:
        products = list_reduziert

    # Verfügbar filtern
    list_verfuegbar = verfuegbar()
    if len(products) > 0:
        products = list(set(products).intersection(list_verfuegbar))
    else:
        products = list_verfuegbar

    # Bio filtern
    list_bio = bio()
    if len(products) > 0:
        products = list(set(products).intersection(list_bio))
    else:
        products = list_bio

    if len(var_searchresults.get()) > 0:
        products = list(set(products).intersection(var_searchresults.get()))

    products = sort(products, int(var_sortierennach.get()))

    for artikelnummer in products:
        for prod in list_products:
            if prod.artikelnummer == artikelnummer:
                prod.setid(overview_canvas.create_window((x, y), window=prod.getproductframe(), anchor='nw'))
                break
        x += 253
        if x == 1268:
            x = 3
            y += 430
    overview_canvas.config(scrollregion=overview_canvas.bbox("all"))




##############################
# Hotbar (Logo, Suchbar, Account, Warenkorb)


# Suchbar
searchbar = Entry(hotbar, width=42, font=("Calibri", 20), textvariable=var_searchinput, borderwidth=0)
searchbar.place(x=324, y=45)

searchnotfound = Label(hotbar, width=18, font=("Calibri", 10), fg='red', bg='white', textvariable=var_searchnotfound, borderwidth=0)
searchnotfound.place(x=560, y=94)

def enter_clicked(self):
    searchbutton_clicked()

searchbar.bind('<Return>', enter_clicked)

# Suchknopf
searchbutton = Button(hotbar, width=34, height=34, image=img_search, command=searchbutton_clicked, borderwidth=0)
searchbutton.photo = img_search
searchbutton.place(x=922, y=45)

# Kontoinformationen
vorname = StringVar(master=window, value='')
nachname = StringVar(master=window, value='')
email = StringVar(master=window, value='')
telefon = StringVar(master=window, value='')
strasse = StringVar(master=window, value='')
plz = StringVar(master=window, value='')
ort = StringVar(master=window, value='')

# Kontoinformationen anzeigen
def accountbutton_clicked():
    # Konto Fesnter
    account = Tk()
    account.title("Konto")
    account.resizable(0, 0)

    # Konto Fenster Images
    img_account_background = PhotoImage(file='account_background.gif', master=account)
    img_save = PhotoImage(file='save.gif', master=account)

    # Konto Frame
    account_frame = Frame(account, width=300, height=400, bg='white')

    # Konto Hintergrund
    account_background = Label(account_frame, image=img_account_background, bg='white', borderwidth=0)
    account_background.photo = img_account_background
    account_background.place(x=0, y=0)

    # Konto Eingabe Vorname
    var_vorname = StringVar(master=account, value=vorname.get())
    account_vorname = Entry(account_frame, width=16, font=("Calibri", 12), textvariable=var_vorname, borderwidth=0)
    account_vorname.place(x=140, y=14)

    # Konto Eingabe Nachname
    var_nachname = StringVar(master=account, value=nachname.get())
    account_nachname = Entry(account_frame, width=16, font=("Calibri", 12), textvariable=var_nachname, borderwidth=0)
    account_nachname.place(x=140, y=64)

    # Konto Eingabe E-Mail
    var_email = StringVar(master=account, value=email.get())
    account_email = Entry(account_frame, width=22, font=("Calibri", 8), textvariable=var_email, borderwidth=0)
    account_email.place(x=140, y=118)

    # Konto Eingabe Telefon
    var_telefon = StringVar(master=account, value=telefon.get())
    account_telefon = Entry(account_frame, width=16, font=("Calibri", 12), textvariable=var_telefon, borderwidth=0)
    account_telefon.place(x=140, y=164)

    # Konto Eingabe Straße
    var_strasse = StringVar(master=account, value=strasse.get())
    account_strasse = Entry(account_frame, width=22, font=("Calibri", 8), textvariable=var_strasse, borderwidth=0)
    account_strasse.place(x=140, y=218)

    # Konto Eingabe PLZ
    var_plz = StringVar(master=account, value=plz.get())
    account_plz = Entry(account_frame, width=16, font=("Calibri", 12), textvariable=var_plz, borderwidth=0)
    account_plz.place(x=140, y=264)

    # Konto Eingabe Ort
    var_ort = StringVar(master=account, value=ort.get())
    account_ort = Entry(account_frame, width=16, font=("Calibri", 12), textvariable=var_ort, borderwidth=0)
    account_ort.place(x=140, y=314)

    def savebutton_clicked():
        vorname.set(str(var_vorname.get()))
        nachname.set(str(var_nachname.get()))
        email.set(str(var_email.get()))
        telefon.set(str(var_telefon.get()))
        strasse.set(str(var_strasse.get()))
        plz.set(str(var_plz.get()))
        ort.set(str(var_ort.get()))
        account.destroy()

    # Kontoinformationen abspeichern
    savebutton = Button(account_frame, image=img_save, command=savebutton_clicked ,bg='white', borderwidth=0)
    savebutton.photo = img_save
    savebutton.place(x=265, y=365)

    # Packen
    account_frame.pack()
    account.mainloop()

# Account
accountbutton = Button(hotbar, image=img_account, bg='white', command=accountbutton_clicked, borderwidth=0)
accountbutton.photo = img_account
accountbutton.place(x=1078, y=38)

# Warenkorb anzeigen
def shoppingbagbutton_clicked():
    shoppingbag = Tk()
    shoppingbag.title("Warenkorb")
    shoppingbag.resizable(0, 0)

    # Variablen
    shoppingbag_products = []
    var_summe = StringVar(master=shoppingbag, value=0)
    var_mwst = StringVar(master=shoppingbag, value=0)

    def updateSumme():
        summe = 0.0
        query = {'in_shoppingbag': {'$gt': 0}}
        doc = db.col.find(query)
        for entry in doc:
            summe += float((float(entry['preis']) - float(entry['rabatt']) + float(entry['pfand'])) * int(entry['in_shoppingbag']))
        mwst = 0.19 * summe
        var_summe.set(str("%.2f" % float(summe)) + ' €')
        var_mwst.set(str("%.2f" % float(mwst)) + ' €')

    updateSumme()

    # Warenkorb Fenster Images
    img_shoppingbag_background = PhotoImage(file='shoppingbag_background.gif', master=shoppingbag)
    img_export = PhotoImage(file="export.gif", master=shoppingbag)

    # Warenkorb Frame
    shoppingbag_frame = Frame(shoppingbag, width=660, height=800, bg='white')

    # Warenkorb Produkt Übersicht
    def shoppingbag_on_mousewheel(event):
        shoppingbag_overview_canvas.yview_scroll(-1 * int(event.delta / 120), "units")

    shoppingbag_overview = Frame(shoppingbag_frame, width=660, height=650, bg='white')
    shoppingbag_overview.grid(rows=1, columns=1)
    shoppingbag_overview.grid_rowconfigure(0, weight=1)
    shoppingbag_overview.grid_columnconfigure(0, weight=1)
    shoppingbag_overview.grid_propagate(False)

    shoppingbag_overview_canvas = Canvas(shoppingbag_overview, width=660, height=650, bg="white")
    shoppingbag_overview_canvas.grid(row=0, column=0)
    shoppingbag_overview_canvas.bind_all("<MouseWheel>", shoppingbag_on_mousewheel)

    shoppingbag_overview_scrollbar = Scrollbar(shoppingbag_overview, orient="vertical", command=shoppingbag_overview_canvas.yview)
    shoppingbag_overview_scrollbar.grid(row=0, column=1, sticky='ns')

    shoppingbag_overview_canvas.configure(yscrollcommand=shoppingbag_overview_scrollbar.set)

    shoppingbag_overview_frame = Frame(shoppingbag_overview_canvas, bg="white")
    shoppingbag_overview_frame.grid(row=0, column=0)

    shoppingbag_overview_canvas.create_window((0, 0), window=shoppingbag_overview_frame, anchor='nw')

    # Produkte anzeigen
    def createShoppingbagProducts():
        x, y = 20, 20
        query = {'in_shoppingbag': {'$gt': 0}}
        doc = db.col.find(query)
        for entry in doc:
            artikelnummer = int(entry['artikelnummer'])
            prod = ShoppingbagProduct(shoppingbag, x, y, shoppingbag_overview_canvas, db, artikelnummer, var_summe, var_mwst)
            shoppingbag_products.append(prod)
            y += 220
        shoppingbag_overview_canvas.config(scrollregion=shoppingbag_overview_canvas.bbox("all"))

    createShoppingbagProducts()

    # Summe Hintergrund
    shoppingbag_summe_frame = Frame(shoppingbag_frame, width=660, height=150)
    shoppingbag_summe_frame.grid(row=1, column=0)

    shoppingbag_background = Label(shoppingbag_summe_frame, image=img_shoppingbag_background, bg='white', borderwidth=0)
    shoppingbag_background.photo = img_shoppingbag_background
    shoppingbag_background.place(x=0, y=0)

    # Export
    def exportbutton_clicked():
        PDF(db, shoppingbag, filedialog.askdirectory(), vorname, nachname, email, telefon, strasse, plz, ort, list_products, shoppingbag_instance)

    exportbutton = Button(shoppingbag_summe_frame, image=img_export, command=exportbutton_clicked, bg='white', highlightbackground='white', borderwidth=0)
    exportbutton.photo = img_export
    exportbutton.place(x=20, y=108)

    # Summe
    shoppingbag_summe = Label(shoppingbag_summe_frame, width=10, height=1, font=("Calibri", 18, 'bold'), bg='white', highlightbackground='white', textvariable=var_summe, borderwidth=0, anchor='center')
    shoppingbag_summe.place(x=510, y=39)

    # Mehrwertsteuer
    shoppingbag_mwst = Label(shoppingbag_summe_frame, width=10, height=1, font=("Calibri", 12), bg='white', highlightbackground='white', textvariable=var_mwst, borderwidth=0, anchor='center')
    shoppingbag_mwst.place(x=534, y=101)

    # Packen
    shoppingbag_overview_canvas.config(scrollregion=shoppingbag_overview_canvas.bbox("all"))
    shoppingbag_frame.pack()
    shoppingbag.mainloop()

# Warenkorb
shoppingbagbutton.config(command=shoppingbagbutton_clicked)
shoppingbagbutton.place(x=1203, y=38)

shoppingbag_background = Label(hotbar, image=img_shoppingbag, bg='white', borderwidth=0)
shoppingbag_background.photo = img_shoppingbag
shoppingbag_background.place(x=1213, y=74)
shoppingbaglabel = Label(hotbar, width=4, height=1, font=("Calibri", 12), bg=color_lightgrey, textvariable=var_shoppingbag, borderwidth=0, anchor='center')
shoppingbaglabel.place(x=1216, y=76)




##############################
# Filter

# Hintergrund
filter_background = Label(filter, image=img_filter, bg='white', borderwidth=0)
filter_background.photo = img_filter
filter_background.place(x=0, y=0)

# Filter Sortieren-Nach Knopf
def sortierennachbutton_clicked():
    showProducts()

sortierennachbutton = Menubutton(filter, image=img_filter_sortierennach, borderwidth=0)
sortierennachbutton.photo = img_filter_sortierennach
sortierennachmenu = Menu(sortierennachbutton, tearoff=0, bg='white')
sortierennachbutton['menu'] = sortierennachmenu
sortierennachmenu.add_radiobutton(label='Alphabetisch (A-Z)', command=sortierennachbutton_clicked, variable=var_sortierennach, value=0)
sortierennachmenu.add_radiobutton(label='Alphabetisch (Z-A)', command=sortierennachbutton_clicked, variable=var_sortierennach, value=1)
sortierennachmenu.add_radiobutton(label='Beliebtheit', command=sortierennachbutton_clicked, variable=var_sortierennach, value=2)
sortierennachmenu.add_radiobutton(label='Preis absteigend', command=sortierennachbutton_clicked, variable=var_sortierennach, value=3)
sortierennachmenu.add_radiobutton(label='Preis aufsteigend', command=sortierennachbutton_clicked, variable=var_sortierennach, value=4)
sortierennachbutton.place(x=21, y=29)

# Filter Preis Knopf
def preisbutton_clicked():
    showProducts()

preisbutton = Menubutton(filter, image=img_filter_preis, borderwidth=0)
preisbutton.photo = img_filter_preis
preismenu = Menu(preisbutton, tearoff=0, bg='white')
preisbutton['menu'] = preismenu
preismenu.add_checkbutton(label='bis 0,99€', command=preisbutton_clicked, variable=var_0_1)
preismenu.add_checkbutton(label='1,00€ - 1,99€', command=preisbutton_clicked, variable=var_1_2)
preismenu.add_checkbutton(label='2,00€ - 4,99€', command=preisbutton_clicked, variable=var_2_5)
preismenu.add_checkbutton(label='ab 5,00€', command=preisbutton_clicked, variable=var_5_inf)
preisbutton.place(x=182, y=29)

# Filter Kategorie Knopf
def kategoriebutton_clicked():
    showProducts()

kategoriebutton = Menubutton(filter, image=img_filter_kategorie, borderwidth=0)
kategoriebutton.photo = img_filter_kategorie
kategoriemenu = Menu(kategoriebutton, tearoff=0, bg='white')
kategoriebutton['menu'] = kategoriemenu
kategoriemenu.add_checkbutton(label='Backwaren', command=kategoriebutton_clicked, variable=var_backwaren)
kategoriemenu.add_checkbutton(label='Drogerieartikel', command=kategoriebutton_clicked, variable=var_drogerieartikel)
kategoriemenu.add_checkbutton(label='Eier und Milchprodukte', command=kategoriebutton_clicked, variable=var_eierundmilchprodukte)
kategoriemenu.add_checkbutton(label='Fisch, Fleisch und Wurst', command=kategoriebutton_clicked, variable=var_fischfleischundwurst)
kategoriemenu.add_checkbutton(label='Gemüse', command=kategoriebutton_clicked, variable=var_gemuese)
kategoriemenu.add_checkbutton(label='Getränke', command=kategoriebutton_clicked, variable=var_getraenke)
kategoriemenu.add_checkbutton(label='Obst', command=kategoriebutton_clicked, variable=var_obst)
kategoriemenu.add_checkbutton(label='Süßwaren', command=kategoriebutton_clicked, variable=var_suesswaren)
kategoriemenu.add_checkbutton(label='Teigwaren', command=kategoriebutton_clicked, variable=var_teigwaren)
kategoriemenu.add_checkbutton(label='Tiefkühlprodukte', command=kategoriebutton_clicked, variable=var_tiefkuehlprodukte)
kategoriebutton.place(x=342, y=29)

# Filter Reduziert
def reduziertbutton_clicked():
    showProducts()

reduziertbutton = Menubutton(filter, image=img_filter_reduziert, borderwidth=0)
reduziertbutton.photo = img_filter_reduziert
reduziertmenu = Menu(reduziertbutton, tearoff=0, bg='white')
reduziertbutton['menu'] = reduziertmenu
reduziertmenu.add_checkbutton(label='Reduziert', command=reduziertbutton_clicked, variable=var_reduziert)
reduziertbutton.place(x=500, y=29)

# Filter Verfügbar Knopf
def verfuegbarbutton_clicked():
    showProducts()

verfuegbarbutton = Menubutton(filter, image=img_filter_verfuegbar, borderwidth=0)
verfuegbarbutton.photo = img_filter_verfuegbar
verfuegbarmenu = Menu(verfuegbarbutton, tearoff=0, bg='white')
verfuegbarbutton['menu'] = verfuegbarmenu
verfuegbarmenu.add_checkbutton(label='Verfügbar', command=verfuegbarbutton_clicked, variable=var_verfuegbar)
verfuegbarbutton.place(x=664, y=29)

# Filter Bio Knopf
def biobutton_clicked():
    showProducts()

biobutton = Menubutton(filter, image=img_filter_bio, borderwidth=0)
biobutton.photo = img_filter_bio
biomenu = Menu(biobutton, tearoff=0, bg='white')
biobutton['menu'] = biomenu
biomenu.add_checkbutton(label='Bio', command=biobutton_clicked, variable=var_bio)
biobutton.place(x=821, y=29)

# Filter zurücksetzen
def resetbutton_clicked():
    var_sortierennach.set(0)
    var_0_1.set(0)
    var_1_2.set(0)
    var_2_5.set(0)
    var_5_inf.set(0)
    var_backwaren.set(0)
    var_drogerieartikel.set(0)
    var_eierundmilchprodukte.set(0)
    var_fischfleischundwurst.set(0)
    var_gemuese.set(0)
    var_getraenke.set(0)
    var_obst.set(0)
    var_suesswaren.set(0)
    var_teigwaren.set(0)
    var_tiefkuehlprodukte.set(0)
    var_reduziert.set(0)
    var_verfuegbar.set(0)
    var_bio.set(0)
    showProducts()

# Filter Reset Knopf
resetbutton = Button(filter, image=img_filter_reset, command=resetbutton_clicked, borderwidth=0)
resetbutton.photo = img_filter_reset
resetbutton.place(x=1030, y=30)




##############################
# Angebot

# Hintergrund
angebot_background = Label(angebot, image=img_angebot, bg='white', borderwidth=0)
angebot_background.photo = img_angebot
angebot_background.place(x=0, y=0)

# Angebot Bild
angebot_bild = Label(angebot, image=img_blank, bg='white', borderwidth=0)
angebot_bild.photo = img_blank
angebot_bild.place(x=402, y=22)

# Angebot Name
var_angebot_name = StringVar(master=window, value='')
angebot_name = Label(angebot, width=26, height=1, font=("Calibri", 12), bg='white', highlightbackground='white', textvariable=var_angebot_name, borderwidth=0, anchor='center')
angebot_name.place(x=660, y=56)

# Angebot Preis
var_angebot_preis = StringVar(master=window, value='')
angebot_preis = Label(angebot, width=6, height=1, font=("Calibri", 10), bg='white', highlightbackground='white', textvariable=var_angebot_preis, borderwidth=0, anchor='center')
angebot_preis.place(x=740, y=94)

# Angebot Linie
angebot_line = Label(angebot, image=img_line, bg='white', borderwidth=0)
angebot_line.photo = img_line
angebot_line.place(x=742, y=102)

# Angebot Reduzierter Preis
var_angebot_reduzierterpreis = StringVar(master=window, value='')
angebot_reduzierterpreis = Label(angebot, width=6, height=1, font=("Calibri", 10), fg='red', bg='white', highlightbackground='white', textvariable=var_angebot_reduzierterpreis, borderwidth=0, anchor='center')
angebot_reduzierterpreis.place(x=786, y=94)

# Reduzierte Produkte
angebot_bilder = []
angebot_namen = []
angebot_preise = []
angebot_reduzierterpreise = []

query = {'rabatt': {'$gt': 0}}
doc = db.col.find(query)
for entry in doc:
    angebot_bilder.append(str(entry['imgsmall']))
    angebot_namen.append(str(entry['name']))
    angebot_preise.append(str("%.2f" % float(entry['preis']) + '€'))
    angebot_reduzierterpreise.append(str("%.2f" % float(float(entry['preis']) - float(entry['rabatt']))) + '€')

# Anfangs Angebot anzeigen
var_index = IntVar(master=window, value=0)

def setAngebot():
    angebot_anfangsbild = PhotoImage(file=angebot_bilder[var_index.get()])
    angebot_bild.config(image=angebot_anfangsbild)
    angebot_bild.photo = angebot_anfangsbild
    var_angebot_name.set(angebot_namen[var_index.get()])
    var_angebot_preis.set(angebot_preise[var_index.get()])
    var_angebot_reduzierterpreis.set(angebot_reduzierterpreise[var_index.get()])

setAngebot()

# Vorheriges Angebot anzeigen
def previousbutton_clicked():
    var_index.set(var_index.get() - 1)
    if var_index.get() == -1:
        var_index.set(len(angebot_namen) - 1)
    elif var_index.get() == len(angebot_namen):
        var_index.set(0)
    setAngebot()

# Nächstes Angebot anzeigen
def nextbutton_clicked():
    var_index.set(var_index.get() + 1)
    if var_index.get() == -1:
        var_index.set(len(angebot_namen) - 1)
    elif var_index.get() == len(angebot_namen):
        var_index.set(0)
    setAngebot()

# Knöpfe
previousbutton = Button(angebot, image=img_previous, command=previousbutton_clicked, borderwidth=0)
previousbutton.photo = img_previous
previousbutton.place(x=372, y=74)
nextbutton = Button(angebot, image=img_next, command=nextbutton_clicked, borderwidth=0)
nextbutton.photo = img_next
nextbutton.place(x=896, y=74)




##############################
# Packen
overview_canvas.config(scrollregion=overview_canvas.bbox("all"))
frame.pack()
window.mainloop()
